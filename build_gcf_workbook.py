import os
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.worksheet.datavalidation import DataValidation
from openpyxl.formatting.rule import FormulaRule
from openpyxl.utils import get_column_letter

OUT = "/home/user/workspace/egc/artifacts/EVEglyphDesign_GCF_Scoring_Workbook.xlsx"
os.makedirs(os.path.dirname(OUT), exist_ok=True)

CREAM = "FDFAF4"
CREAM2 = "F7F2E7"
INK = "1A1A1A"
LINE = "E7E1D3"
MUTE = "6B665C"
ORANGE = "E87722"

F = "Calibri"
fill_cream = PatternFill("solid", start_color=CREAM, end_color=CREAM)
fill_cream2 = PatternFill("solid", start_color=CREAM2, end_color=CREAM2)
fill_orange = PatternFill("solid", start_color=ORANGE, end_color=ORANGE)

thin = Side(style="thin", color=LINE)
box = Border(left=thin, right=thin, top=thin, bottom=thin)
orange_bottom = Border(bottom=Side(style="medium", color=ORANGE))

f_title = Font(name=F, size=16, bold=True, color=INK)
f_sub = Font(name=F, size=10, color=MUTE)
f_h2 = Font(name=F, size=12, bold=True, color=INK)
f_head = Font(name=F, size=10, bold=True, color=INK)
f_body = Font(name=F, size=10, color=INK)
f_body_b = Font(name=F, size=10, bold=True, color=INK)
f_mute = Font(name=F, size=10, color=MUTE)
f_big = Font(name=F, size=20, bold=True, color=INK)
f_orange_b = Font(name=F, size=11, bold=True, color=ORANGE)

top_wrap = Alignment(horizontal="left", vertical="top", wrap_text=True)
ctr = Alignment(horizontal="center", vertical="center", wrap_text=True)
head_al = Alignment(horizontal="left", vertical="center", wrap_text=True)

wb = openpyxl.Workbook()


def paint(ws, max_row, max_col):
    ws.sheet_view.showGridLines = False
    for r in range(1, max_row + 1):
        for c in range(1, max_col + 1):
            cell = ws.cell(row=r, column=c)
            if cell.fill.fgColor.rgb in (None, "00000000") or cell.fill.fill_type is None:
                cell.fill = fill_cream
            if cell.font is None or cell.font.name != F:
                pass
    ws.sheet_properties.tabColor = ORANGE


def widths(ws, mapping):
    for col, w in mapping.items():
        ws.column_dimensions[col].width = w


# ---------------------------------------------------------------- Read me
ws = wb.active
ws.title = "Read me"
widths(ws, {"A": 3, "B": 30, "C": 104})
ws["B2"] = "EVEglyphDesign — Global Compliance Framework Assessment"
ws["B2"].font = f_title
ws.merge_cells("B2:C2")
ws["B3"] = "EgD-GCF-001 · Vendor AI proposal scoring workbook"
ws["B3"].font = f_sub
ws.merge_cells("B3:C3")
ws["B4"] = None
for c in ("B", "C"):
    ws[f"{c}4"].border = orange_bottom

readme_rows = [
    ("What this is",
     "A working instrument for scoring a vendor's AI proposal against the eight structural-dependency domains "
     "of the EgD Global Compliance Framework (EgD-GCF-001). It is designed to be completed by a consultant, "
     "sitting with the proposal and the draft contract, before the client signs."),
    ("How to use it",
     "1. Read the 'Domain detail' sheet and put the listed questions to the vendor in writing.\n"
     "2. Score each of the eight domains on the 'Scoring sheet' using the dropdown in the Score column.\n"
     "3. Record the evidence for every score — a URL or a specific contract clause reference.\n"
     "4. Answer the four veto questions on the 'Result' sheet, then read the verdict."),
    ("The evidence standard",
     "2 — EVIDENCED. A contractual commitment or a testable artifact. A clause you can point to by number, "
     "or an export you have actually executed.\n"
     "1 — ASSERTED. The vendor says so. Marketing copy, a security questionnaire answer, or a trust-centre page. "
     "A trust-centre page caps the score at 1, however comprehensive it appears.\n"
     "0 — ABSENT. No commitment and no assertion, or the vendor declined to answer."),
    ("Evidence is mandatory",
     "Every score must carry an entry in 'Evidence cited'. A score without a citation is not a score; it is an "
     "impression. Cite the URL, or the clause number and the agreement it sits in (e.g. 'MSA §7.3', 'DPA Annex II')."),
    ("Reading the result",
     "The 'Result' sheet totals the eight domains out of 16 and returns a verdict band. Four veto conditions sit "
     "above the total. Any one of them returns DECLINE regardless of the score, because no amount of strength "
     "elsewhere compensates for a structural loss of control."),
    ("Not legal advice",
     "This is a decision-support instrument. It does not constitute legal advice and does not substitute for "
     "review by qualified counsel in the relevant jurisdictions. The scores record what the consultant found; "
     "the decision remains the client's."),
    ("Licence",
     "May be used with clients unmodified, with attribution to EVEglyphDesign. May not be rebranded."),
]

r = 6
for head, body in readme_rows:
    ws.cell(row=r, column=2, value=head).font = f_h2
    ws.cell(row=r, column=2).alignment = top_wrap
    c = ws.cell(row=r, column=3, value=body)
    c.font = f_body
    c.alignment = top_wrap
    nlines = sum(max(1, (len(seg) // 100) + 1) for seg in body.split("\n"))
    ws.row_dimensions[r].height = max(18, 14 * nlines + 4)
    ws.cell(row=r, column=2).border = Border(bottom=thin)
    ws.cell(row=r, column=3).border = Border(bottom=thin)
    r += 1

ws.freeze_panes = "A5"
paint(ws, 40, 6)

# ---------------------------------------------------------------- Scoring sheet
sc = wb.create_sheet("Scoring sheet")
widths(sc, {"A": 3, "B": 10, "C": 30, "D": 52, "E": 9, "F": 34, "G": 34})
sc["B2"] = "Scoring sheet"
sc["B2"].font = f_title
sc.merge_cells("B2:G2")
sc["B3"] = "Score each domain 0 / 1 / 2. Every score requires a cited URL or contract clause."
sc["B3"].font = f_sub
sc.merge_cells("B3:G3")

HDR = 5
headers = ["Domain ID", "Control domain", "What evidenced (2) looks like", "Score (0/1/2)",
           "Evidence cited (URL or clause)", "Notes"]
for i, h in enumerate(headers):
    c = sc.cell(row=HDR, column=2 + i, value=h)
    c.font = f_head
    c.fill = fill_cream2
    c.alignment = head_al
    c.border = Border(left=thin, right=thin, top=thin, bottom=Side(style="medium", color=ORANGE))
sc.row_dimensions[HDR].height = 32

domains = [
    ("SD-1", "Derived-data title",
     "Who owns embeddings, fine-tunes and extracted structure derived from the customer's data. "
     "Evidenced = the contract assigns title in derived artifacts to the customer."),
    ("SD-2", "Purpose limitation and training use",
     "Whether customer data or derived artifacts train models beyond the customer's own tenant. "
     "Evidenced = a contractual prohibition, not a policy page."),
    ("SD-3", "Extractability",
     "A complete, timestamped, full-text export the customer can execute without vendor assistance "
     "(the FTTE standard). Evidenced = a documented, tested export path."),
    ("SD-4", "Switching and functional equivalence",
     "Switching assistance obligations, a transition window stated in days, and charges no greater than "
     "directly incurred cost."),
    ("SD-5", "Residency, transfer and lawful access",
     "Where data rests AND where inference runs, plus the lawful-access exposure that follows from the answer."),
    ("SD-6", "Substrate and subprocessor transparency",
     "Which model, running on whose infrastructure, and notice to the customer before that changes."),
    ("SD-7", "Access, logging and audit",
     "Customer-exportable logs including vendor-personnel access, and an attestation whose scope actually "
     "covers the AI service rather than the surrounding platform."),
    ("SD-8", "Exit, deletion and survivability",
     "Deletion that is verifiable, and an explicit statement of what survives termination."),
]

FIRST = HDR + 1
for j, (did, name, desc) in enumerate(domains):
    r = FIRST + j
    sc.cell(row=r, column=2, value=did).font = f_body_b
    sc.cell(row=r, column=2).alignment = ctr
    sc.cell(row=r, column=3, value=name).font = f_body_b
    sc.cell(row=r, column=3).alignment = top_wrap
    sc.cell(row=r, column=4, value=desc).font = f_body
    sc.cell(row=r, column=4).alignment = top_wrap
    sc.cell(row=r, column=5).alignment = ctr
    sc.cell(row=r, column=5).font = f_body_b
    sc.cell(row=r, column=6).alignment = top_wrap
    sc.cell(row=r, column=6).font = f_body
    sc.cell(row=r, column=7).alignment = top_wrap
    sc.cell(row=r, column=7).font = f_body
    for col in range(2, 8):
        cell = sc.cell(row=r, column=col)
        cell.border = box
        cell.fill = fill_cream2 if col == 5 else fill_cream
    sc.row_dimensions[r].height = 48

LAST = FIRST + len(domains) - 1
dv = DataValidation(type="list", formula1='"0,1,2"', allow_blank=True, showDropDown=False)
dv.error = "Score must be 0, 1 or 2."
dv.errorTitle = "Invalid score"
dv.prompt = "0 = absent, 1 = asserted, 2 = evidenced"
dv.promptTitle = "Score"
sc.add_data_validation(dv)
dv.add(f"E{FIRST}:E{LAST}")

sc.cell(row=LAST + 2, column=3, value="Total (calculated on the Result sheet)").font = f_mute
sc.cell(row=LAST + 2, column=5, value=f"=SUM(E{FIRST}:E{LAST})").font = f_body_b
sc.cell(row=LAST + 2, column=5).alignment = ctr
sc.cell(row=LAST + 2, column=5).border = Border(top=Side(style="medium", color=ORANGE))

sc.freeze_panes = "A6"
paint(sc, LAST + 6, 8)

# ---------------------------------------------------------------- Result
rs = wb.create_sheet("Result")
widths(rs, {"A": 3, "B": 62, "C": 20, "D": 44})
rs["B2"] = "Result"
rs["B2"].font = f_title
rs.merge_cells("B2:D2")
rs["B3"] = "Computed from the scoring sheet. No manual entry below except the four veto answers."
rs["B3"].font = f_sub
rs.merge_cells("B3:D3")

rs["B5"] = "Veto conditions"
rs["B5"].font = f_h2
rs["C5"] = "Yes / No"
rs["C5"].font = f_head
rs["C5"].alignment = ctr
rs["D5"] = "Why it vetoes"
rs["D5"].font = f_head
for col in ("B", "C", "D"):
    rs[f"{col}5"].fill = fill_cream2
    rs[f"{col}5"].border = Border(left=thin, right=thin, top=thin, bottom=Side(style="medium", color=ORANGE))
rs.row_dimensions[5].height = 24

vetoes = [
    ("1. The vendor takes title to derived artifacts (embeddings, fine-tunes, extracted structure).",
     "The customer's own data returns to it as someone else's asset."),
    ("2. The vendor holds unbounded rights to train on customer data.",
     "Confidentiality and derived value both leave the tenant, irreversibly."),
    ("3. There is no independent extraction path the customer can execute alone.",
     "Exit depends on the goodwill of the party being exited."),
    ("4. The inference jurisdiction is undisclosed.",
     "Lawful-access exposure cannot be assessed at all, let alone accepted."),
]
VF = 6
for j, (q, why) in enumerate(vetoes):
    r = VF + j
    rs.cell(row=r, column=2, value=q).font = f_body
    rs.cell(row=r, column=2).alignment = top_wrap
    rs.cell(row=r, column=3, value="No").font = f_body_b
    rs.cell(row=r, column=3).alignment = ctr
    rs.cell(row=r, column=3).fill = fill_cream2
    rs.cell(row=r, column=4, value=why).font = f_mute
    rs.cell(row=r, column=4).alignment = top_wrap
    for col in range(2, 5):
        rs.cell(row=r, column=col).border = box
    rs.row_dimensions[r].height = 32
VL = VF + len(vetoes) - 1

dv2 = DataValidation(type="list", formula1='"Yes,No"', allow_blank=True, showDropDown=False)
dv2.errorTitle = "Invalid answer"
dv2.error = "Answer Yes or No."
rs.add_data_validation(dv2)
dv2.add(f"C{VF}:C{VL}")

SCORE_RANGE = f"'Scoring sheet'!E{FIRST}:E{LAST}"
r = VL + 2
rs.cell(row=r, column=2, value="Total score (out of 16)").font = f_h2
rs.cell(row=r, column=2).alignment = Alignment(horizontal="left", vertical="center")
tc = rs.cell(row=r, column=3, value=f"=SUM({SCORE_RANGE})")
tc.font = f_big
tc.alignment = ctr
tc.number_format = "0"
rs.cell(row=r, column=4, value="Sum of the eight domain scores. Maximum 16.").font = f_mute
rs.cell(row=r, column=4).alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)
rs.row_dimensions[r].height = 34
TOTAL_CELL = f"C{r}"
for col in range(2, 5):
    rs.cell(row=r, column=col).border = box
    rs.cell(row=r, column=col).fill = fill_cream2 if col == 3 else fill_cream

r += 1
rs.cell(row=r, column=2, value="Domains still unscored").font = f_h2
rs.cell(row=r, column=2).alignment = Alignment(horizontal="left", vertical="center")
uc = rs.cell(row=r, column=3, value=f"=COUNTBLANK({SCORE_RANGE})")
uc.font = f_big
uc.alignment = ctr
uc.number_format = "0"
rs.cell(row=r, column=4, value="A verdict read while this is above zero is provisional.").font = f_mute
rs.cell(row=r, column=4).alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)
rs.row_dimensions[r].height = 34
for col in range(2, 5):
    rs.cell(row=r, column=col).border = box
    rs.cell(row=r, column=col).fill = fill_cream2 if col == 3 else fill_cream

r += 2
VERDICT_ROW = r
rs.cell(row=r, column=2, value="Verdict").font = f_h2
rs.cell(row=r, column=2).alignment = Alignment(horizontal="left", vertical="center")
verdict = (
    f'=IF(COUNTIF(C{VF}:C{VL},"Yes")>0,"DECLINE \u2014 veto condition triggered",'
    f'IF(COUNTBLANK({SCORE_RANGE})=8,"Not yet scored",'
    f'IF({TOTAL_CELL}>=14,"Proceed",'
    f'IF({TOTAL_CELL}>=10,"Proceed with conditions",'
    f'IF({TOTAL_CELL}>=6,"Renegotiate before signature","Decline")))))'
)
vc = rs.cell(row=r, column=3, value=verdict)
vc.font = Font(name=F, size=14, bold=True, color=INK)
vc.alignment = ctr
rs.merge_cells(start_row=r, start_column=3, end_row=r, end_column=4)
rs.row_dimensions[r].height = 40
for col in range(2, 5):
    rs.cell(row=r, column=col).border = Border(left=thin, right=thin,
                                               top=Side(style="medium", color=ORANGE),
                                               bottom=Side(style="medium", color=ORANGE))
rs.cell(row=r, column=3).fill = fill_cream2

veto_flag = f'COUNTIF($C${VF}:$C${VL},"Yes")>0'
rs.conditional_formatting.add(
    f"C{VERDICT_ROW}:D{VERDICT_ROW}",
    FormulaRule(formula=[veto_flag], fill=fill_orange,
                font=Font(name=F, size=14, bold=True, color=CREAM), stopIfTrue=True))
rs.conditional_formatting.add(
    f"C{VF}:C{VL}",
    FormulaRule(formula=['$C6="Yes"'], fill=fill_orange,
                font=Font(name=F, size=10, bold=True, color=CREAM), stopIfTrue=True))

r += 2
rs.cell(row=r, column=2, value="Verdict bands").font = f_h2
r += 1
bands = [
    ("14 - 16", "Proceed"),
    ("10 - 13", "Proceed with conditions"),
    ("6 - 9", "Renegotiate before signature"),
    ("0 - 5", "Decline"),
    ("Any veto = Yes", "DECLINE \u2014 veto condition triggered (overrides the total)"),
]
for band, label in bands:
    rs.cell(row=r, column=2, value=band).font = f_body_b
    rs.cell(row=r, column=2).alignment = Alignment(horizontal="left", vertical="center")
    c = rs.cell(row=r, column=3, value=label)
    c.font = f_orange_b if band.startswith("Any") else f_body
    c.alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)
    rs.merge_cells(start_row=r, start_column=3, end_row=r, end_column=4)
    for col in range(2, 5):
        rs.cell(row=r, column=col).border = box
    rs.row_dimensions[r].height = 20
    r += 1

rs.freeze_panes = "A5"
paint(rs, r + 4, 6)

# ---------------------------------------------------------------- Domain detail
dd = wb.create_sheet("Domain detail")
widths(dd, {"A": 3, "B": 26, "C": 104})
dd["B2"] = "Domain detail"
dd["B2"].font = f_title
dd.merge_cells("B2:C2")
dd["B3"] = "The fuller description of each domain, with the questions to put to the vendor in writing."
dd["B3"].font = f_sub
dd.merge_cells("B3:C3")

detail = [
    ("SD-1", "Derived-data title",
     "Who owns the embeddings, fine-tunes, indexes and extracted structure derived from the customer's data. "
     "Raw data ownership is rarely contested; the derived layer is where title quietly transfers. Evidenced (2) "
     "means the contract assigns title in derived artifacts to the customer, by name.",
     ["Does the agreement assign title in embeddings, vector indexes, fine-tuned weights and extracted structure "
      "to us? Quote the clause.",
      "If title does not pass to us, what licence do we hold in those artifacts, and does it survive termination?",
      "Can we obtain the derived artifacts themselves on exit, or only the source data we supplied?"]),
    ("SD-2", "Purpose limitation and training use",
     "Whether customer data, or artifacts derived from it, are used to train, tune or evaluate models beyond the "
     "customer's own tenant. Evidenced (2) means a contractual prohibition. A policy page, however clearly written, "
     "is an assertion and caps at 1.",
     ["Is there a contractual prohibition — not a policy statement — on using our data or derived artifacts to "
      "train or tune any model outside our tenant?",
      "Does that prohibition bind your subprocessors and your model providers, and how is it flowed down?",
      "What is used for abuse monitoring or evaluation, who sees it, and for how long is it retained?"]),
    ("SD-3", "Extractability",
     "Whether the customer can obtain a complete, timestamped, full-text export it can execute without vendor "
     "assistance — the FTTE standard. Evidenced (2) means a documented export path that has actually been tested, "
     "not a support ticket you can raise.",
     ["Can we execute a complete full-text, timestamped export ourselves, without opening a ticket? Point us to "
      "the documentation.",
      "What is in the export and what is left behind — attachments, metadata, audit history, derived artifacts?",
      "Will you support a test export during the pilot, and will you contract to maintain that path?"]),
    ("SD-4", "Switching and functional equivalence",
     "The obligations that apply when the customer leaves: switching assistance, a transition window stated in "
     "days, and charges no greater than the cost directly incurred by the vendor. Evidenced (2) means all three "
     "appear as obligations, with numbers.",
     ["What switching assistance are you obliged to provide, and over what transition window, stated in days?",
      "Are switching charges limited to costs directly incurred by you? Show us the clause and the rate basis.",
      "In what format is the data delivered, and is it usable by a competing service without bespoke conversion?"]),
    ("SD-5", "Residency, transfer and lawful access",
     "Where data rests and — separately — where inference runs. The two answers often differ, and the second one "
     "determines the lawful-access exposure. Evidenced (2) means both are contractually fixed, with the transfer "
     "mechanism and access exposure named.",
     ["Name the jurisdictions where our data rests, and separately the jurisdictions where inference runs.",
      "Which foreign lawful-access regimes could reach our data or our prompts as a result, and through which entity?",
      "Are the residency commitments contractual and change-controlled, or configuration we could lose in an update?"]),
    ("SD-6", "Substrate and subprocessor transparency",
     "Which model is running, on whose infrastructure, and what notice the customer receives before that changes. "
     "Evidenced (2) means the model and the hosting substrate are named, with a contractual notice period before "
     "substitution.",
     ["Name the model and version behind the service, and the infrastructure provider it runs on.",
      "What notice do we receive before the model, the version or the substrate changes, and can we object?",
      "Where is the current subprocessor list published, and are we notified of additions before they take effect?"]),
    ("SD-7", "Access, logging and audit",
     "Whether the customer can export its own logs — including records of vendor-personnel access — and whether "
     "any third-party attestation genuinely covers the AI service rather than the surrounding platform. Evidenced "
     "(2) means exportable logs plus an attestation whose scope statement names the AI service.",
     ["Can we export our own access and activity logs, including every instance of vendor-personnel access to our "
      "tenant?",
      "Show us the attestation's scope statement: does it name the AI service, or only the platform around it?",
      "How are privileged vendor-side actions authorised, recorded and made visible to us in near real time?"]),
    ("SD-8", "Exit, deletion and survivability",
     "Whether deletion is verifiable and what survives termination — backups, derived artifacts, aggregated "
     "statistics, model improvements. Evidenced (2) means a deletion obligation with a certificate or equivalent "
     "verification, and an explicit list of what persists.",
     ["On termination, what exactly is deleted, on what timetable, and do we receive a certificate of deletion?",
      "What survives — backups, embeddings, fine-tunes, aggregate statistics, model improvements — and for how long?",
      "How can we verify deletion independently rather than accepting your confirmation?"]),
]

from openpyxl.worksheet.pagebreak import Break
block_starts = {}
r = 5
for did, name, desc, qs in detail:
    block_starts[did] = r
    hc = dd.cell(row=r, column=2, value=f"{did}  {name}")
    hc.font = Font(name=F, size=12, bold=True, color=INK)
    hc.alignment = Alignment(horizontal="left", vertical="center")
    dd.merge_cells(start_row=r, start_column=2, end_row=r, end_column=3)
    dd.cell(row=r, column=2).fill = fill_cream2
    dd.cell(row=r, column=3).fill = fill_cream2
    for col in (2, 3):
        dd.cell(row=r, column=col).border = Border(top=thin, bottom=Side(style="medium", color=ORANGE),
                                                   left=thin, right=thin)
    dd.row_dimensions[r].height = 24
    r += 1

    dd.cell(row=r, column=2, value="What the domain covers").font = f_body_b
    dd.cell(row=r, column=2).alignment = top_wrap
    dc = dd.cell(row=r, column=3, value=desc)
    dc.font = f_body
    dc.alignment = top_wrap
    dd.row_dimensions[r].height = 14 * ((len(desc) // 100) + 1) + 6
    for col in (2, 3):
        dd.cell(row=r, column=col).border = Border(left=thin, right=thin, bottom=thin)
    r += 1

    for i, q in enumerate(qs):
        dd.cell(row=r, column=2, value="Ask the vendor, in writing" if i == 0 else "").font = f_body_b
        dd.cell(row=r, column=2).alignment = top_wrap
        qc = dd.cell(row=r, column=3, value=f"{i + 1}.  {q}")
        qc.font = f_body
        qc.alignment = top_wrap
        dd.row_dimensions[r].height = 15 * ((len(q) // 115) + 1) + 4
        for col in (2, 3):
            dd.cell(row=r, column=col).border = Border(left=thin, right=thin,
                                                       bottom=thin if i == len(qs) - 1 else None)
        r += 1
    r += 1

for did in ("SD-4", "SD-7"):
    dd.row_breaks.append(Break(id=block_starts[did] - 1))

dd.freeze_panes = "A5"
paint(dd, r + 2, 5)

# ---------------------------------------------------------------- print setup
for sheet in wb.worksheets:
    sheet.page_setup.orientation = "landscape"
    sheet.page_setup.fitToWidth = 1
    sheet.page_setup.fitToHeight = 0
    sheet.sheet_properties.pageSetUpPr.fitToPage = True
    sheet.print_options.horizontalCentered = False

wb.save(OUT)
print("saved", OUT)
